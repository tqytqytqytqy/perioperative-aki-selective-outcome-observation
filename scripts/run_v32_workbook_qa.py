#!/usr/bin/env python3
"""Validate the standard-dependency v3.2 aggregate workbook against all CSV sources."""

from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from build_v32_workbook_public import OUTPUT, SHEETS, public_disclosure_frame


ROOT = Path(__file__).resolve().parents[1]


def equivalent(left, right) -> bool:
    if pd.isna(left) and right is None:
        return True
    if isinstance(left, (int, float, np.integer, np.floating)) and isinstance(right, (int, float)):
        return bool(np.isclose(float(left), float(right), rtol=0, atol=1e-10, equal_nan=True))
    if isinstance(left, str) and isinstance(right, (int, float, np.integer, np.floating)):
        try:
            return bool(np.isclose(float(left), float(right), rtol=0, atol=1e-10))
        except ValueError:
            pass
    if isinstance(right, str) and isinstance(left, (int, float, np.integer, np.floating)):
        try:
            return bool(np.isclose(float(left), float(right), rtol=0, atol=1e-10))
        except ValueError:
            pass
    return str(left) == str(right)


def main() -> int:
    rows: list[dict[str, str]] = []

    def add(check: str, passed: bool, detail: object) -> None:
        rows.append({"check": check, "status": "pass" if passed else "fail", "detail": str(detail)})

    workbook = load_workbook(OUTPUT, read_only=False, data_only=False)
    relative_output = str(OUTPUT.relative_to(ROOT))
    add("workbook_opens", True, relative_output)
    add("sheet_count_30", len(workbook.sheetnames) == 30, len(workbook.sheetnames))
    expected_names = ["README", "QA_Summary"] + [name for name, _ in SHEETS]
    add("sheet_order", workbook.sheetnames == expected_names, workbook.sheetnames)

    formula_cells = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")
    add("readme_formula_links_present", len(formula_cells) >= 8, len(formula_cells))
    add("no_broken_formula_references", not any("#REF!" in value for value in formula_cells), formula_cells)

    for sheet_name, filename in SHEETS:
        frame = public_disclosure_frame(filename, pd.read_csv(ROOT / "tables" / filename))
        sheet = workbook[sheet_name]
        add(f"{sheet_name}::dimensions", sheet.max_row == len(frame) + 1 and sheet.max_column == len(frame.columns), f"{sheet.max_row}x{sheet.max_column}")
        headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
        add(f"{sheet_name}::headers", headers == frame.columns.tolist(), headers)
        values_match = True
        mismatch = ""
        for row_index, values in enumerate(frame.itertuples(index=False, name=None), 2):
            for column_index, source_value in enumerate(values, 1):
                workbook_value = sheet.cell(row_index, column_index).value
                if not equivalent(source_value, workbook_value):
                    values_match = False
                    mismatch = f"row={row_index}; col={column_index}; csv={source_value}; xlsx={workbook_value}"
                    break
            if not values_match:
                break
        add(f"{sheet_name}::all_values", values_match, mismatch or f"{len(frame)} rows")
        filter_present = bool(sheet.auto_filter.ref) or bool(sheet.tables)
        add(
            f"{sheet_name}::filter_navigation_present",
            filter_present,
            f"freeze={sheet.freeze_panes}; auto_filter={sheet.auto_filter.ref}; table_filters={len(sheet.tables)}; freeze panes are an advisory UI feature",
        )
        styled_headers = all(cell.font.bold and cell.fill.fill_type == "solid" for cell in sheet[1])
        add(f"{sheet_name}::header_style", styled_headers, "bold solid-fill headers")

    result = pd.DataFrame(rows)
    result.to_csv(ROOT / "qa/reproducible_workbook_checks_v32.csv", index=False)
    failed = result.loc[result["status"].eq("fail")]
    summary = {
        "status": "pass" if failed.empty else "fail",
        "workbook": relative_output,
        "checks": len(result),
        "passed": int(result["status"].eq("pass").sum()),
        "failed": int(len(failed)),
        "failed_checks": failed["check"].tolist(),
    }
    (ROOT / "qa/reproducible_workbook_summary_v32.json").write_text(
        json.dumps(summary, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps(summary, indent=2))
    return 0 if failed.empty else 1


if __name__ == "__main__":
    raise SystemExit(main())
