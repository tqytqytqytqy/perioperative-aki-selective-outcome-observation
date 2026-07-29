#!/usr/bin/env python3
"""Validate the journal-neutral v3.2.5 release metadata and public boundary."""

from __future__ import annotations

import importlib.util
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TITLE = (
    "Selective Outcome Observation Across Development, Recalibration, and "
    "Temporal Evaluation of a Transported Perioperative AKI Model"
)
VERSION = "v3.2.5"
RELEASE_DATE = "2026-07-29"
CONCEPT_DOI = "10.5281/zenodo.21366088"
EXPECTED_AUTHORS = [
    "Qingyu Teng",
    "Tingting Niu",
    "Qian Chen",
    "Min Tao",
    "Ziyan Gu",
    "Qi Li",
    "Yingya Zhao",
    "Hui Zhang",
]
EXPECTED_ZENODO_NAMES = [
    "Teng, Qingyu",
    "Niu, Tingting",
    "Chen, Qian",
    "Tao, Min",
    "Gu, Ziyan",
    "Li, Qi",
    "Zhao, Yingya",
    "Zhang, Hui",
]
STALE_AUTHOR_NAMES = (
    "Junde Han",
    "Han, Junde",
    "Jin Zhao",
    "Zhao, Jin",
    "Tao Xu",
    "Xu, Tao",
)
PROHIBITED_SUFFIXES = {
    ".docx",
    ".doc",
    ".parquet",
    ".feather",
    ".pkl",
    ".pickle",
    ".rds",
    ".sav",
    ".dta",
}


def add(
    checks: list[dict[str, object]],
    name: str,
    passed: bool,
    detail: object,
) -> None:
    checks.append({"check": name, "passed": bool(passed), "detail": detail})


def load_author_module():
    path = ROOT / "scripts" / "author_metadata_v32.py"
    spec = importlib.util.spec_from_file_location("author_metadata_v32", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot import {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def main() -> int:
    checks: list[dict[str, object]] = []

    zenodo = json.loads((ROOT / ".zenodo.json").read_text(encoding="utf-8"))
    citation_text = (ROOT / "CITATION.cff").read_text(encoding="utf-8")
    readme = (ROOT / "README.md").read_text(encoding="utf-8")
    model_card = (ROOT / "MODEL_CARD.md").read_text(encoding="utf-8")
    license_code = (ROOT / "LICENSE-CODE").read_text(encoding="utf-8")
    author_module = load_author_module()

    add(checks, "zenodo_title", zenodo.get("title") == TITLE, zenodo.get("title"))
    add(checks, "zenodo_version", zenodo.get("version") == VERSION, zenodo.get("version"))
    add(
        checks,
        "zenodo_release_date",
        zenodo.get("publication_date") == RELEASE_DATE,
        zenodo.get("publication_date"),
    )
    creator_names = [creator.get("name") for creator in zenodo.get("creators", [])]
    add(checks, "zenodo_creator_order", creator_names == EXPECTED_ZENODO_NAMES, creator_names)
    add(checks, "zenodo_open_access", zenodo.get("access_right") == "open", zenodo.get("access_right"))
    add(checks, "zenodo_license", zenodo.get("license") == "MIT", zenodo.get("license"))
    add(
        checks,
        "zenodo_related_release",
        any(
            item.get("identifier", "").endswith("/tree/v3.2.5")
            for item in zenodo.get("related_identifiers", [])
        ),
        zenodo.get("related_identifiers", []),
    )

    citation_title = re.search(r"^title:\s*(.+)$", citation_text, re.M)
    citation_version = re.search(r"^version:\s*(.+)$", citation_text, re.M)
    citation_date = re.search(r"^date-released:\s*(.+)$", citation_text, re.M)
    citation_doi = re.search(r"^doi:\s*(.+)$", citation_text, re.M)
    citation_url = re.search(r"^url:\s*(.+)$", citation_text, re.M)
    add(
        checks,
        "citation_title",
        bool(citation_title and citation_title.group(1) == TITLE),
        citation_title.group(1) if citation_title else None,
    )
    add(
        checks,
        "citation_version",
        bool(citation_version and citation_version.group(1) == "3.2.5"),
        citation_version.group(1) if citation_version else None,
    )
    add(
        checks,
        "citation_release_date",
        bool(citation_date and citation_date.group(1) == RELEASE_DATE),
        citation_date.group(1) if citation_date else None,
    )
    add(
        checks,
        "citation_unpublished_version_doi_absent",
        citation_doi is None,
        citation_doi.group(1) if citation_doi else None,
    )
    add(
        checks,
        "citation_unpublished_version_doi_url_absent",
        citation_url is None,
        citation_url.group(1) if citation_url else None,
    )
    identifier_values = re.findall(r"^\s+value:\s*(.+)$", citation_text, re.M)
    add(checks, "citation_concept_doi", CONCEPT_DOI in identifier_values, identifier_values)
    citation_authors = [
        f"{given.strip()} {family.strip()}"
        for family, given in re.findall(
            r"^\s+- family-names:\s*(.+)\n\s+given-names:\s*(.+)$",
            citation_text,
            re.M,
        )
    ]
    add(checks, "citation_author_order", citation_authors == EXPECTED_AUTHORS, citation_authors)

    local_authors = [author["name"] for author in author_module.AUTHOR_METADATA]
    add(checks, "author_module_order", local_authors == EXPECTED_AUTHORS, local_authors)
    add(checks, "author_module_count", len(local_authors) == 8, len(local_authors))
    add(
        checks,
        "equal_contributors",
        author_module.equal_contributors()
        == ["Qingyu Teng", "Tingting Niu", "Qian Chen"],
        author_module.equal_contributors(),
    )
    corresponding = [
        author["name"] for author in author_module.corresponding_authors()
    ]
    add(
        checks,
        "corresponding_authors",
        corresponding == ["Qi Li", "Yingya Zhao", "Hui Zhang"],
        corresponding,
    )

    expected_license_line = (
        "Copyright (c) 2026 Qingyu Teng, Tingting Niu, Qian Chen, Min Tao, "
        "Ziyan Gu, Qi Li, Yingya Zhao, and Hui Zhang"
    )
    add(checks, "license_holders", expected_license_line in license_code, expected_license_line)
    add(checks, "readme_version", readme.startswith("# Selective outcome observation study v3.2.5"), readme.splitlines()[0])
    add(checks, "readme_unpublished_doi_absent", "10.5281/zenodo.21663368" not in readme, "10.5281/zenodo.21663368")
    add(checks, "model_card_version", model_card.startswith("# Model card v3.2.5"), model_card.splitlines()[0])
    add(checks, "model_card_unpublished_doi_absent", "10.5281/zenodo.21663368" not in model_card, "10.5281/zenodo.21663368")

    all_paths = [path for path in ROOT.rglob("*") if path.is_file()]
    prohibited = [
        str(path.relative_to(ROOT))
        for path in all_paths
        if path.suffix.lower() in PROHIBITED_SUFFIXES
        or path.name == "runtime_config_v32.local.json"
    ]
    add(checks, "public_boundary_no_restricted_files", not prohibited, prohibited)
    journal_bound_paths = [
        str(path.relative_to(ROOT))
        for path in all_paths
        if re.search(
            r"(?:jamia|bja|jama[_ -]?network[_ -]?open)",
            str(path.relative_to(ROOT)),
            re.I,
        )
    ]
    add(checks, "journal_neutral_paths", not journal_bound_paths, journal_bound_paths)

    metadata_text = "\n".join(
        (
            citation_text,
            readme,
            model_card,
            license_code,
            json.dumps(zenodo, ensure_ascii=False),
        )
    )
    stale_names = [name for name in STALE_AUTHOR_NAMES if name in metadata_text]
    add(checks, "no_stale_authors", not stale_names, stale_names)
    stale_versions = [
        value
        for value in (
            "10.5281/zenodo.21447516",
            "10.5281/zenodo.21663368",
        )
        if value in metadata_text
    ]
    add(
        checks,
        "no_stale_release_identifier",
        "10.5281/zenodo.21447516" not in metadata_text
        and "10.5281/zenodo.21663368" not in metadata_text
        and zenodo.get("version") == VERSION
        and str(citation_version.group(1) if citation_version else "") == "3.2.5",
        {
            "hits": stale_versions,
            "zenodo_version": zenodo.get("version"),
            "citation_version": citation_version.group(1) if citation_version else None,
        },
    )
    targeted_journal_hits = re.findall(
        r"\b(?:JAMIA|BJA|JAMA Network Open)\b",
        metadata_text,
        flags=re.I,
    )
    add(checks, "journal_neutral_metadata", not targeted_journal_hits, targeted_journal_hits)

    workbook_path = (
        ROOT / "workbook" / "AKI_selective_outcome_v32_all_tables_reproducible.xlsx"
    )
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    add(checks, "workbook_sheet_count", len(workbook.sheetnames) == 30, len(workbook.sheetnames))
    add(
        checks,
        "workbook_release_heading",
        workbook["README"]["A1"].value
        == "Selective outcome observation study v3.2.5",
        workbook["README"]["A1"].value,
    )
    workbook.close()

    failures = [check for check in checks if not check["passed"]]
    output = {
        "status": "pass" if not failures else "fail",
        "checks": len(checks),
        "passed": len(checks) - len(failures),
        "failed": len(failures),
        "failures": failures,
        "details": checks,
    }
    output_path = ROOT / "qa" / "release_metadata_checks_v325.json"
    output_path.write_text(
        json.dumps(output, indent=2, ensure_ascii=True) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {key: output[key] for key in ("status", "checks", "passed", "failed")},
            indent=2,
        )
    )
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
