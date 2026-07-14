#!/usr/bin/env python3
"""Build the aggregate v3.2 workbook with standard openpyxl dependencies."""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "workbook" / "AKI_selective_outcome_v32_all_tables_reproducible.xlsx"

SHEETS = [
    ("01_CohortFlow", "01_cohort_flow_and_roles_v32.csv"),
    ("27_Characteristics", "27_source_update_observation_flow_v32.csv"),
    ("28_ObsDiagnostics", "28_stage_observation_diagnostics_v32.csv"),
    ("29_UpdateEstimands", "29_mover2021_recalibration_estimands_v32.csv"),
    ("30_TargetStrategies", "30_update_selection_target_performance_v32.csv"),
    ("31_Canonical", "31_canonical_primary_bootstrap_v32.csv"),
    ("32_Balance", "32_stage_observation_balance_v32.csv"),
    ("33_Missingness", "33_predictor_missingness_v32.csv"),
    ("34_WorkloadStrategies", "34_update_strategy_workload_v32.csv"),
    ("35_MainCohorts", "35_main_table1_cohorts_v32.csv"),
    ("36_MainPropagation", "36_main_table2_propagation_v32.csv"),
    ("37_MainPerformance", "37_main_table3_canonical_performance_v32.csv"),
    ("37_SelectionChain", "37_selection_chain_primary_comparators_v32.csv"),
    ("38_MainWorkload", "38_main_table4_workload_v32.csv"),
    ("38_UpdateMNAR", "38_update_mnar_propagation_v32.csv"),
    ("39_MetricMap", "39_metric_reporting_map_v32.csv"),
    ("40_TargetMNAR", "40_target_mnar_sensitivity_v32.csv"),
    ("41_TargetBounds", "41_target_nonparametric_bounds_v32.csv"),
    ("42_MNARAnchors", "42_mnar_anchor_comparison_v32.csv"),
    ("43_SourceMNAR", "43_source_mnar_propagation_v32.csv"),
    ("44_AuxOutcome", "44_auxiliary_outcome_model_diagnostics_v32.csv"),
    ("45_SourcePreprocess", "45_source_preprocessing_audit_v32.csv"),
    ("46_Truncation", "46_weight_truncation_sensitivity_v32.csv"),
    ("47_MNARIntervals", "47_mnar_chain_bootstrap_intervals_v32.csv"),
    ("48_VariableMap", "48_raw_to_analysis_variable_map_v32.csv"),
    ("49_VitalDB", "49_vitaldb_supportive_v32.csv"),
    ("50_INSPIRECrLevels", "50_inspire_creatinine_release_levels_v32.csv"),
    ("51_MOVERCoarsening", "51_mover_deterministic_coarsening_v32.csv"),
]

NAVY = "17365D"
BLUE = "DCE6F1"
PALE = "EFF4F8"
WHITE = "FFFFFF"
GRID = Side(style="thin", color="D9E2F3")


def write_frame(ws, frame: pd.DataFrame) -> None:
    for column, name in enumerate(frame.columns, 1):
        cell = ws.cell(1, column, str(name))
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), 2):
        for column, value in enumerate(values, 1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row_index, column, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=GRID)
            if isinstance(value, float):
                cell.number_format = "0.0000"
        if row_index % 2 == 0:
            for cell in ws[row_index]:
                cell.fill = PatternFill("solid", fgColor=PALE)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    for column in range(1, len(frame.columns) + 1):
        values = [str(frame.columns[column - 1])]
        values.extend(str(value) for value in frame.iloc[:200, column - 1].dropna())
        width = min(42, max(10, max(len(value) for value in values) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width


def metric_row(frame: pd.DataFrame, metric: str) -> int:
    matches = frame.index[frame["metric"].eq(metric)].tolist()
    if len(matches) != 1:
        raise ValueError(f"Expected one {metric} row, found {len(matches)}")
    return matches[0] + 2


def main() -> int:
    tables = {name: pd.read_csv(ROOT / "tables" / filename) for name, filename in SHEETS}
    qa = json.loads((ROOT / "qa" / "v32_statistical_qa_summary.json").read_text())
    rebuild = pd.read_csv(ROOT / "qa" / "raw_rebuild_equivalence_v32.csv")
    canonical = tables["31_Canonical"]

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Selective outcome observation study v3.2"
    ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:D1")
    readme = [
        ("Analysis status", "Scientific analysis complete; submission NO-GO pending human gates"),
        ("Study type", "Post-exploration methodological audit; not a clinical deployment study"),
        ("Canonical source", "data/processed/canonical_bootstrap_distribution_v32.parquet"),
        ("MNAR interval source", "data/processed/mnar_chain_bootstrap_v32.parquet"),
        ("Statistical QA", "=QA_Summary!B2"),
        ("Target eligible n", "='01_CohortFlow'!C4"),
        ("Canonical O/E", f"='31_Canonical'!E{metric_row(canonical, 'oe_ratio')}"),
        ("Canonical slope", f"='31_Canonical'!E{metric_row(canonical, 'calibration_slope')}"),
        ("Canonical AUROC", f"='31_Canonical'!E{metric_row(canonical, 'auroc')}"),
        ("Alerts at 5%", "='38_MainWorkload'!B2"),
        ("Alerts at 10%", "='38_MainWorkload'!B3"),
        ("Alerts at 20%", "='38_MainWorkload'!B4"),
        ("Release boundary", "Raw and patient-level derived data are excluded from public release"),
    ]
    for row, (label, value) in enumerate(readme, 3):
        ws.cell(row, 1, label).font = Font(bold=True, color=NAVY)
        ws.cell(row, 2, value)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    qa_ws = wb.create_sheet("QA_Summary")
    qa_rows = [
        ("Item", "Result"),
        ("Statistical QA", f"{qa['status'].upper()} ({qa['passed']}/{qa['checks']})"),
        ("Raw rebuild differing cells", int(rebuild["differing_cells"].sum())),
        ("Bootstrap replicates", 1000),
        ("MNAR bootstrap rows", 18000),
        ("Submission status", "NO-GO pending author, ethics, disclosure, and independent signoff"),
    ]
    for row, values in enumerate(qa_rows, 1):
        for column, value in enumerate(values, 1):
            qa_ws.cell(row, column, value)
    for cell in qa_ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    qa_ws.column_dimensions["A"].width = 34
    qa_ws.column_dimensions["B"].width = 68
    qa_ws.sheet_view.showGridLines = False

    for name, _ in SHEETS:
        write_frame(wb.create_sheet(name), tables[name])

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Built {OUTPUT} with {len(wb.sheetnames)} sheets")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
