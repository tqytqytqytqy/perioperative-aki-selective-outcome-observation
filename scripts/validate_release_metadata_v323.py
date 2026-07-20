#!/usr/bin/env python3
"""Validate the journal-neutral v3.2.3 release metadata and public boundary."""

from __future__ import annotations

import argparse
import importlib.util
import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TITLE = (
    "Missing Postoperative Creatinine Outcomes Across the Development, "
    "Recalibration, and Evaluation of a Perioperative Acute Kidney Injury Model"
)
VERSION = "v3.2.3"
RELEASE_DATE = "2026-07-20"
CONCEPT_DOI = "10.5281/zenodo.21366088"
EXPECTED_AUTHORS = [
    "Qingyu Teng",
    "Yingya Zhao",
    "Jin Zhao",
    "Qian Chen",
    "Min Tao",
    "Qi Li",
    "Tao Xu",
    "Hui Zhang",
]
EXPECTED_ZENODO_NAMES = [
    "Teng, Qingyu",
    "Zhao, Yingya",
    "Zhao, Jin",
    "Chen, Qian",
    "Tao, Min",
    "Li, Qi",
    "Xu, Tao",
    "Zhang, Hui",
]
SUPERSEDED_AUTHOR_NAMES = ("Junde Han", "Han, Junde")
PROHIBITED_SUFFIXES = {
    ".docx",
    ".doc",
    ".parquet",
    ".feather",
    ".pkl",
    ".pickle",
    ".rds",
    ".sav",
    ".dta",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--require-doi", action="store_true")
    return parser.parse_args()


def add(checks: list[dict[str, object]], name: str, passed: bool, detail: str) -> None:
    checks.append({"check": name, "passed": bool(passed), "detail": detail})


def load_author_module():
    path = ROOT / "scripts" / "author_metadata_v32.py"
    spec = importlib.util.spec_from_file_location("author_metadata_v32", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot import {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def main() -> int:
    args = parse_args()
    checks: list[dict[str, object]] = []

    zenodo = json.loads((ROOT / ".zenodo.json").read_text(encoding="utf-8"))
    citation = (ROOT / "CITATION.cff").read_text(encoding="utf-8")
    readme = (ROOT / "README.md").read_text(encoding="utf-8")
    model_card = (ROOT / "MODEL_CARD.md").read_text(encoding="utf-8")
    license_code = (ROOT / "LICENSE-CODE").read_text(encoding="utf-8")
    author_module = load_author_module()

    add(checks, "zenodo_title", zenodo.get("title") == TITLE, str(zenodo.get("title")))
    add(checks, "zenodo_version", zenodo.get("version") == VERSION, str(zenodo.get("version")))
    add(
        checks,
        "zenodo_release_date",
        zenodo.get("publication_date") == RELEASE_DATE,
        str(zenodo.get("publication_date")),
    )
    creator_names = [creator.get("name") for creator in zenodo.get("creators", [])]
    add(checks, "zenodo_creator_order", creator_names == EXPECTED_ZENODO_NAMES, repr(creator_names))
    add(checks, "zenodo_open_access", zenodo.get("access_right") == "open", str(zenodo.get("access_right")))
    add(checks, "zenodo_license", zenodo.get("license") == "MIT", str(zenodo.get("license")))
    add(
        checks,
        "zenodo_related_release",
        any(
            item.get("identifier", "").endswith("/tree/v3.2.3")
            for item in zenodo.get("related_identifiers", [])
        ),
        repr(zenodo.get("related_identifiers", [])),
    )

    add(checks, "citation_title", f"title: {TITLE}" in citation, TITLE)
    add(checks, "citation_version", "version: 3.2.3" in citation, "version: 3.2.3")
    add(checks, "citation_release_date", f"date-released: {RELEASE_DATE}" in citation, RELEASE_DATE)
    add(checks, "citation_concept_doi", CONCEPT_DOI in citation, CONCEPT_DOI)
    for family, given in [
        ("Teng", "Qingyu"),
        ("Zhao", "Yingya"),
        ("Zhao", "Jin"),
        ("Chen", "Qian"),
        ("Tao", "Min"),
        ("Li", "Qi"),
        ("Xu", "Tao"),
        ("Zhang", "Hui"),
    ]:
        add(
            checks,
            f"citation_author_{family}_{given}",
            f"family-names: {family}\n    given-names: {given}" in citation,
            f"{given} {family}",
        )

    local_authors = [author["name"] for author in author_module.AUTHOR_METADATA]
    add(checks, "author_module_order", local_authors == EXPECTED_AUTHORS, repr(local_authors))
    add(checks, "author_module_count", len(local_authors) == 8, str(len(local_authors)))
    add(
        checks,
        "equal_contributors",
        author_module.equal_contributors() == ["Qingyu Teng", "Yingya Zhao"],
        repr(author_module.equal_contributors()),
    )
    add(
        checks,
        "corresponding_authors",
        [author["name"] for author in author_module.corresponding_authors()]
        == ["Qi Li", "Tao Xu", "Hui Zhang"],
        repr([author["name"] for author in author_module.corresponding_authors()]),
    )

    expected_license_line = (
        "Copyright (c) 2026 Qingyu Teng, Yingya Zhao, Jin Zhao, Qian Chen, "
        "Min Tao, Qi Li, Tao Xu, and Hui Zhang"
    )
    add(checks, "license_holders", expected_license_line in license_code, expected_license_line)
    add(checks, "readme_version", "study v3.2.3" in readme, readme.splitlines()[0])
    add(checks, "model_card_version", model_card.startswith("# Model card v3.2.3"), model_card.splitlines()[0])

    all_paths = [path for path in ROOT.rglob("*") if path.is_file()]
    prohibited = [
        str(path.relative_to(ROOT))
        for path in all_paths
        if path.suffix.lower() in PROHIBITED_SUFFIXES
        or path.name == "runtime_config_v32.local.json"
    ]
    add(checks, "public_boundary_no_restricted_files", not prohibited, repr(prohibited))
    journal_bound_paths = [
        str(path.relative_to(ROOT))
        for path in all_paths
        if re.search(r"(?:jamia|bja|jama[_ -]?network[_ -]?open)", str(path.relative_to(ROOT)), re.I)
    ]
    add(checks, "journal_neutral_paths", not journal_bound_paths, repr(journal_bound_paths))

    metadata_text = "\n".join((citation, readme, model_card, license_code, json.dumps(zenodo)))
    stale_names = [name for name in SUPERSEDED_AUTHOR_NAMES if name in metadata_text]
    add(checks, "no_superseded_authors", not stale_names, repr(stale_names))

    doi_match = re.search(r"^doi:\s*(10\.5281/zenodo\.\d+)\s*$", citation, re.M)
    doi_value = doi_match.group(1) if doi_match else ""
    if args.require_doi:
        add(checks, "version_doi_reserved", bool(doi_value), doi_value or "missing")
        add(checks, "version_doi_url", f"url: https://doi.org/{doi_value}" in citation, doi_value)
        add(checks, "readme_version_doi", doi_value in readme, doi_value)
        add(checks, "model_card_version_doi", doi_value in model_card, doi_value)
    else:
        pending_ok = doi_value or "TO_BE_RESERVED_BEFORE_RELEASE" in citation
        add(checks, "version_doi_pending_or_valid", bool(pending_ok), doi_value or "pending")

    workbook_path = ROOT / "workbook" / "AKI_selective_outcome_v32_all_tables_reproducible.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    add(checks, "workbook_sheet_count", len(workbook.sheetnames) == 30, str(len(workbook.sheetnames)))
    add(
        checks,
        "workbook_release_heading",
        workbook["README"]["A1"].value == "Missing postoperative creatinine outcomes study v3.2.3",
        str(workbook["README"]["A1"].value),
    )
    workbook.close()

    failures = [check for check in checks if not check["passed"]]
    output = {
        "status": "pass" if not failures else "fail",
        "checks": len(checks),
        "passed": len(checks) - len(failures),
        "failed": len(failures),
        "require_doi": args.require_doi,
        "failures": failures,
        "details": checks,
    }
    output_path = ROOT / "qa" / "release_metadata_checks_v323.json"
    output_path.write_text(json.dumps(output, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    print(json.dumps({key: output[key] for key in ("status", "checks", "passed", "failed")}, indent=2))
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
